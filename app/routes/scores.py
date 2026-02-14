"""Score Entry Routes"""
from flask import Blueprint, render_template, request, jsonify
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'src'))
from database.db_manager import DatabaseManager
from logic.grading import calculate_grade

bp = Blueprint('scores', __name__, url_prefix='/scores')
db = DatabaseManager()

@bp.route('/')
def index():
    """Score entry page"""
    sessions = db.execute_query('SELECT * FROM sessions ORDER BY name DESC')
    classes = db.execute_query('SELECT * FROM classes ORDER BY level, stream')
    # Subjects are loaded dynamically based on class selection via JS
    return render_template('scores/index.html', sessions=sessions, classes=classes)

@bp.route('/api/terms/<int:session_id>')
def get_terms(session_id):
    """Get terms for a session"""
    terms = db.execute_query('SELECT * FROM terms WHERE session_id = ? ORDER BY term_number', (session_id,))
    return jsonify([dict(t) for t in terms])

@bp.route('/api/students_in_class/<int:class_id>')
def get_students_in_class(class_id):
    """Get all active students in a class"""
    print(f"Fetching students for class_id: {class_id}")
    # Single line query to ensure matching
    students = db.execute_query(
        'SELECT id, first_name, last_name, reg_number FROM students WHERE class_id = ? AND active_status = 1 ORDER BY last_name, first_name', 
        (class_id,)
    )
    print(f"Found {len(students)} students")
    return jsonify([dict(s) for s in students])

@bp.route('/api/student_score_grid')
def get_student_score_grid():
    """Get score grid for a single student (Rows=Subjects)"""
    student_id = request.args.get('student_id', type=int)
    class_id = request.args.get('class_id', type=int)
    session_id = request.args.get('session_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    if not all([student_id, class_id, session_id, term_id]):
        return jsonify([])
        
    # 1. Get all subjects assigned to this class
    class_subjects = db.execute_query('''
        SELECT s.id, s.name, s.code 
        FROM subjects s
        JOIN class_subjects cs ON s.id = cs.subject_id
        WHERE cs.class_id = ?
        ORDER BY cs.display_order, s.name
    ''', (class_id,))
    
    result = []
    for sub in class_subjects:
        # 2. Get existing score for this student/subject
        score = db.execute_query('''
            SELECT * FROM scores 
            WHERE student_id = ? AND subject_id = ? AND session_id = ? AND term_id = ?
        ''', (student_id, sub['id'], session_id, term_id))
        
        score_data = score[0] if score else {}
        
        result.append({
            'subject_id': sub['id'],
            'subject_name': sub['name'],
            'ca': score_data.get('ca_score', 0),
            'exam': score_data.get('exam_score', 0),
            'total': score_data.get('total', 0),
            'grade': score_data.get('grade', '-'),
            'position': score_data.get('position', '-'),
            'highest': score_data.get('class_highest', '-'),
            'lowest': score_data.get('class_lowest', '-'),
            'average': score_data.get('class_average', '-')
        })
        
    return jsonify(result)

@bp.route('/api/subject_score_grid')
def get_subject_score_grid():
    """Get score grid for all students in a class for one subject (Rows=Students)"""
    class_id = request.args.get('class_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    session_id = request.args.get('session_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    if not all([class_id, subject_id, session_id, term_id]):
        return jsonify([])
        
    # 1. Get all students in this class
    students = db.execute_query('''
        SELECT id, first_name, last_name, reg_number 
        FROM students 
        WHERE class_id = ? AND active_status = 1
        ORDER BY last_name, first_name
    ''', (class_id,))
    
    result = []
    for std in students:
        # 2. Get existing score for this student/subject
        score = db.execute_query('''
            SELECT * FROM scores 
            WHERE student_id = ? AND subject_id = ? AND session_id = ? AND term_id = ?
        ''', (std['id'], subject_id, session_id, term_id))
        
        score_data = score[0] if score else {}
        
        result.append({
            'student_id': std['id'],
            'full_name': f"{std['last_name']} {std['first_name']}",
            'reg_number': std['reg_number'],
            'ca': score_data.get('ca_score', 0),
            'exam': score_data.get('exam_score', 0),
            'total': score_data.get('total', 0),
            'grade': score_data.get('grade', '-'),
            'position': score_data.get('position', '-'),
            'highest': score_data.get('class_highest', '-'),
            'lowest': score_data.get('class_lowest', '-'),
            'average': score_data.get('class_average', '-')
        })
        
    return jsonify(result)

@bp.route('/api/save_subject_scores', methods=['POST'])
def save_subject_scores():
    """Save scores for all students in a class for one subject"""
    data = request.json
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    session_id = data.get('session_id')
    term_id = data.get('term_id')
    scores_data = data.get('scores', [])

    if not all([class_id, subject_id, session_id, term_id]):
        return jsonify({'success': False, 'message': 'Missing required fields'}), 400

    try:
        saved_count = 0
        for entry in scores_data:
            student_id = entry['student_id']
            ca = float(entry['ca'])
            exam = float(entry['exam'])
            total = ca + exam
            grade = calculate_grade(total)

            db.execute_update('''
                INSERT OR REPLACE INTO scores 
                (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, subject_id, session_id, term_id, ca, exam, total, grade))
            saved_count += 1

        # Trigger Ranking
        ranking_message = ""
        try:
            from logic.ranking import RankingEngine
            engine = RankingEngine()
            res = engine.process_class_results(class_id, session_id, term_id)
            if not res['success']:
                print(f"Ranking warning: {res['message']}")
                ranking_message = f" (Ranking warning: {res['message']})"
        except Exception as re:
            print(f"Ranking failed: {re}")
            ranking_message = " (Ranking failed to update)"

        return jsonify({'success': True, 'message': f'Scores saved for {saved_count} students{ranking_message}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/class_subjects/<int:class_id>')
def get_class_subjects(class_id):
    """Get subjects assigned to a class"""
    subjects = db.execute_query('''
        SELECT s.id, s.name, s.code 
        FROM subjects s
        JOIN class_subjects cs ON s.id = cs.subject_id
        WHERE cs.class_id = ?
        ORDER BY s.name
    ''', (class_id,))
    return jsonify([dict(s) for s in subjects])

@bp.route('/api/save_student_scores', methods=['POST'])
def save_student_scores():
    """Save scores for a single student (multiple subjects) with validation"""
    data = request.json
    student_id = data.get('student_id')
    class_id = data.get('class_id')
    session_id = data.get('session_id')
    term_id = data.get('term_id')
    scores_data = data.get('scores', [])

    if not all([student_id, class_id, session_id, term_id]):
        return jsonify({
            'success': False,
            'message': 'Missing required fields'
        }), 400

    if not scores_data:
        return jsonify({
            'success': False,
            'message': 'No scores provided'
        }), 400

    try:
        # Validate all scores first
        for score_entry in scores_data:
            ca_score = float(score_entry['ca'])
            exam_score = float(score_entry['exam'])

            # Range validation
            if not (0 <= ca_score <= 30):
                return jsonify({
                    'success': False,
                    'message': f'CA score must be between 0 and 30 (got {ca_score})'
                }), 400

            if not (0 <= exam_score <= 70):
                return jsonify({
                    'success': False,
                    'message': f'Exam score must be between 0 and 70 (got {exam_score})'
                }), 400

        # Save all scores
        saved_count = 0
        for score_entry in scores_data:
            subject_id = score_entry['subject_id']
            ca_score = float(score_entry['ca'])
            exam_score = float(score_entry['exam'])
            total = ca_score + exam_score
            grade = calculate_grade(total)

            db.execute_update('''
                INSERT OR REPLACE INTO scores
                (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade))
            saved_count += 1

        # Trigger Ranking Calculation
        ranking_success = False
        ranking_msg = ""
        try:
            from logic.ranking import RankingEngine
            engine = RankingEngine()
            res = engine.process_class_results(class_id, session_id, term_id)
            if res['success']:
                ranking_success = True
            else:
                ranking_msg = res['message']
                print(f"Ranking warning: {ranking_msg}")
        except Exception as re_err:
            print(f"Ranking update failed: {re_err}")
            ranking_msg = str(re_err)

        # Fetch updated scores with rankings
        updated_scores = []
        if ranking_success:
            for score_entry in scores_data:
                subject_id = score_entry['subject_id']
                score = db.execute_query('''
                    SELECT * FROM scores
                    WHERE student_id = ? AND subject_id = ? AND session_id = ? AND term_id = ?
                ''', (student_id, subject_id, session_id, term_id))

                if score:
                    updated_scores.append({
                        'subject_id': subject_id,
                        'position': score[0].get('position'),
                        'class_highest': score[0].get('class_highest'),
                        'class_lowest': score[0].get('class_lowest'),
                        'class_average': score[0].get('class_average')
                    })
        
        msg = f'{saved_count} scores saved'
        if not ranking_success:
             msg += f' (Ranking issue: {ranking_msg})'

        return jsonify({
            'success': True,
            'message': msg,
            'updated_scores': updated_scores,
            'ranking_success': ranking_success
        })

    except ValueError as ve:
        return jsonify({
            'success': False,
            'message': f'Invalid score value: {str(ve)}'
        }), 400
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to save scores: {str(e)}'
        }), 500

@bp.route('/api/save', methods=['POST'])
def save_scores():
    """Save scores (Legacy/Subject-Centric)"""
    data = request.json
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    session_id = data.get('session_id')
    term_id = data.get('term_id')
    scores_data = data.get('scores', [])
    
    if not all([class_id, subject_id, session_id, term_id]):
        return jsonify({'success': False, 'message': 'Missing required fields'})
    
    try:
        for score_entry in scores_data:
            student_id = score_entry['student_id']
            ca_score = float(score_entry['ca'])
            exam_score = float(score_entry['exam'])
            total = ca_score + exam_score
            grade = calculate_grade(total)
            
            db.execute_update('''
                INSERT OR REPLACE INTO scores 
                (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade))
        
        return jsonify({'success': True, 'message': 'Scores saved successfully!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# --- Excel Export / Import ---

@bp.route('/api/export_excel')
def export_excel():
    """Export scores to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    from flask import send_file

    try:
        class_id = request.args.get('class_id')
        session_id = request.args.get('session_id')
        term_id = request.args.get('term_id')
        subject_id = request.args.get('subject_id') # Optional (for subject mode)

        if not all([class_id, session_id, term_id]):
            return "Missing parameters", 400

        # Fetch Context
        class_info = db.execute_query('SELECT name FROM classes WHERE id=?', (class_id,))[0]
        session_info = db.execute_query('SELECT name FROM sessions WHERE id=?', (session_id,))[0]
        term_info = db.execute_query('SELECT term_number FROM terms WHERE id=?', (term_id,))[0]
        
        output = BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Scores"

        # Headers
        headers = ['Reg Number', 'Last Name', 'First Name', 'Subject Code', 'Subject Name', 'CA (30)', 'Exam (70)', 'Total']
        sheet.append(headers)
        
        # Style Headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Fetch Data
        if subject_id:
            # Export for specific subject (All students in class)
            subject = db.execute_query('SELECT code, name FROM subjects WHERE id=?', (subject_id,))[0]
            students = db.execute_query('SELECT id, reg_number, last_name, first_name FROM students WHERE class_id=? AND active_status=1 ORDER BY last_name', (class_id,))
            
            for std in students:
                score = db.execute_query('SELECT ca_score, exam_score FROM scores WHERE student_id=? AND subject_id=? AND session_id=? AND term_id=?',
                                        (std['id'], subject_id, session_id, term_id))
                ca = score[0]['ca_score'] if score else 0
                exam = score[0]['exam_score'] if score else 0
                
                sheet.append([
                    std['reg_number'], std['last_name'], std['first_name'], 
                    subject['code'], subject['name'], 
                    ca, exam, ca + exam
                ])
        else:
            # Export all subjects for all students (Big Grid)
            # For simplicity, we'll list rows normally: Student - Subject - Score
            students = db.execute_query('SELECT id, reg_number, last_name, first_name FROM students WHERE class_id=? AND active_status=1 ORDER BY last_name', (class_id,))
            subjects = db.execute_query('SELECT s.id, s.code, s.name FROM subjects s JOIN class_subjects cs ON s.id=cs.subject_id WHERE cs.class_id=? ORDER BY s.name', (class_id,))
            
            for std in students:
                for sub in subjects:
                    score = db.execute_query('SELECT ca_score, exam_score FROM scores WHERE student_id=? AND subject_id=? AND session_id=? AND term_id=?',
                                            (std['id'], sub['id'], session_id, term_id))
                    ca = score[0]['ca_score'] if score else 0
                    exam = score[0]['exam_score'] if score else 0
                    
                    sheet.append([
                        std['reg_number'], std['last_name'], std['first_name'], 
                        sub['code'], sub['name'], 
                        ca, exam, ca + exam
                    ])

        # Auto-adjust column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)
        
        filename = f"Scores_{class_info['name']}_{session_info['name']}_Term{term_info['term_number']}.xlsx".replace('/', '-')
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Export Error: {e}")
        return str(e), 500

@bp.route('/api/import_excel', methods=['POST'])
def import_excel():
    """Import scores from Excel"""
    import openpyxl
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    class_id = request.form.get('class_id')
    session_id = request.form.get('session_id')
    term_id = request.form.get('term_id')
    
    if not all([file, class_id, session_id, term_id]):
         return jsonify({'success': False, 'message': 'Missing parameters'})

    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        
        # Expected Headers: Reg Number | Last Name | First Name | Subject Code | Subject Name | CA | Exam
        # We allow flexible column indices
        headers = {}
        for cell in sheet[1]:
            headers[cell.value.lower().strip()] = cell.col_idx - 1 # 0-based index
            
        required = ['reg number', 'subject code', 'ca (30)', 'exam (70)']
        for req in required:
            if req not in headers:
                 # Try minimal matching
                 if req == 'ca (30)' and 'ca' in headers: headers['ca (30)'] = headers['ca']
                 elif req == 'exam (70)' and 'exam' in headers: headers['exam (70)'] = headers['exam']
                 else:
                    return jsonify({'success': False, 'message': f'Missing column: {req}'})

        success_count = 0
        errors = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                reg_num = row[headers['reg number']]
                sub_code = row[headers['subject code']]
                ca = row[headers['ca (30)']] or 0
                exam = row[headers['exam (70)']] or 0
                
                if not reg_num or not sub_code: continue
                
                # Validation
                try:
                    ca = float(ca)
                    exam = float(exam)
                except:
                    errors.append(f"Row {row_idx}: Invalid number format")
                    continue
                    
                if ca > 30 or exam > 70:
                    errors.append(f"Row {row_idx}: Scores out of range (CA<=30, Exam<=70)")
                    continue

                # Resolve IDs
                std = db.execute_query('SELECT id FROM students WHERE reg_number = ?', (reg_num,))
                if not std:
                    errors.append(f"Row {row_idx}: Student {reg_num} not found")
                    continue
                
                sub = db.execute_query('SELECT id FROM subjects WHERE code = ?', (sub_code,))
                if not sub:
                    errors.append(f"Row {row_idx}: Subject {sub_code} not found")
                    continue
                    
                # Save
                total = ca + exam
                grade = calculate_grade(total)
                
                db.execute_update('''
                    INSERT OR REPLACE INTO scores 
                    (student_id, subject_id, session_id, term_id, ca_score, exam_score, total, grade)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (std[0]['id'], sub[0]['id'], session_id, term_id, ca, exam, total, grade))
                
                success_count += 1
                
            except Exception as row_err:
                errors.append(f"Row {row_idx}: {str(row_err)}")

        # Rankings update
        ranking_msg = ""
        try:
             res = ranking_engine.process_class_results(class_id, session_id, term_id)
        except:
             pass

        return jsonify({
            'success': True,
            'message': f'Imported {success_count} scores. {len(errors)} errors.',
            'errors': errors[:10] # Top 10 errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f"File error: {str(e)}"})
